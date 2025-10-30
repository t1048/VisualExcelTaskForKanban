import datetime as dt
from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from backend.backend import META_ID_COLUMN, TASK_COLUMNS, TaskStore


def _build_initial_workbook(path: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.delete_rows(1, ws.max_row)
    ws.append(TASK_COLUMNS)
    ws.append([
        "未着手",
        "カテゴリA",
        "中分類A",
        "タスクA",
        "Alice",
        "高",
        dt.date(2024, 1, 1),
        "最初のタスク",
    ])
    ws.append([
        "進行中",
        "カテゴリB",
        "中分類B",
        "タスクB",
        "Bob",
        "中",
        dt.date(2024, 1, 2),
        "二番目のタスク",
    ])
    wb.save(path)


def test_bootstrap_persists_ids_and_respects_local_changes(tmp_path):
    excel_path = tmp_path / "board.xlsx"
    sheet_name = "Kanban"
    _build_initial_workbook(excel_path, sheet_name)

    store = TaskStore(excel_path, sheet_name=sheet_name)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    assert META_ID_COLUMN in headers
    id_col_idx = headers.index(META_ID_COLUMN) + 1
    assert ws.column_dimensions[get_column_letter(id_col_idx)].hidden is True
    first_id_cell = ws.cell(row=2, column=id_col_idx).value
    second_id_cell = ws.cell(row=3, column=id_col_idx).value
    assert first_id_cell
    assert second_id_cell

    df = store._df.copy()
    first_row_id = df.iloc[0][META_ID_COLUMN]
    second_row_id = df.iloc[1][META_ID_COLUMN]

    store.update_task(1, {"タスク": "ローカル修正"})
    store.delete_task(2)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    task_col_idx = headers.index("タスク") + 1
    ws.cell(row=2, column=task_col_idx).value = "外部編集"
    ws.cell(row=3, column=task_col_idx).value = "復活してほしくない"
    wb.save(excel_path)

    store.load_excel()

    assert len(store._df) == 1
    assert store._df.iloc[0]["タスク"] == "ローカル修正"
    assert first_row_id in store._dirty_row_ids
    assert second_row_id in store._deleted_row_ids
    assert first_row_id == store._df.iloc[0][META_ID_COLUMN]
