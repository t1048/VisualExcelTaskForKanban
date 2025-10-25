# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import os
import shutil
import threading
from pathlib import Path
from typing import Any, Dict, List
import datetime as dt

import pandas as pd
import webview
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation


REQUIRED_COLUMNS = [
    "No",
    "ステータス",
    "タスク",
    "担当者",
    "優先度",
    "期限",
    "備考",
]
DEFAULT_STATUSES = ["未着手", "進行中", "完了", "保留"]
MAX_VALIDATION_CELLS = 10000


def _to_iso_date_str(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (dt.date, dt.datetime, pd.Timestamp)):
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    try:
        return pd.to_datetime(str(value)).strftime("%Y-%m-%d")
    except Exception:
        return str(value)


def _from_iso_date_str(value: str):
    value = (value or "").strip()
    if not value:
        return pd.NaT
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return pd.NaT


def _normalize_priority(value: Any):
    if value is None or value is pd.NA:
        return pd.NA

    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return pd.NA
        value = trimmed

    if pd.isna(value):
        return pd.NA

    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)

    if pd.isna(num):
        return pd.NA

    if float(num).is_integer():
        return int(num)
    return float(num)


def _format_priority(value: Any):
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float)) and not pd.isna(value):
        if isinstance(value, float) and float(value).is_integer():
            return int(value)
        return value
    return str(value)


class TaskStore:
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self._lock = threading.RLock()
        self._df = pd.DataFrame(columns=REQUIRED_COLUMNS)
        self._statuses: List[str] = list(DEFAULT_STATUSES)
        self._sheet_name: str | None = None
        self._validations: Dict[str, List[str]] = {}
        self.load_excel()

    def load_excel(self):
        with self._lock:
            if not self.excel_path.exists():
                df = pd.DataFrame(columns=REQUIRED_COLUMNS)
                df.to_excel(self.excel_path, index=False)

            wb = load_workbook(self.excel_path, data_only=False)
            sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            self._sheet_name = sheet_name
            self._validations = self._extract_validations(wb, ws)

            df = pd.read_excel(
                self.excel_path,
                dtype={"No": "Int64"},
                sheet_name=sheet_name,
                engine="openpyxl",
            )

            for col in REQUIRED_COLUMNS:
                if col not in df.columns:
                    df[col] = pd.NA

            df = df[REQUIRED_COLUMNS].copy()

            if "期限" in df.columns:
                df["期限"] = pd.to_datetime(df["期限"], errors="coerce").dt.date

            if self._validations.get("ステータス"):
                base = list(self._validations["ステータス"])
                extras = [
                    s
                    for s in df["ステータス"].dropna().astype(str).tolist()
                    if s and s not in base
                ]
                self._statuses = base + extras
            else:
                self._rebuild_statuses_from_df(df)

            self._df = df

    def _rebuild_statuses_from_df(self, df: pd.DataFrame | None = None):
        if df is None:
            df = self._df
        status_values = [
            s for s in df["ステータス"].dropna().astype(str).tolist() if s
        ]
        merged: List[str] = []
        for name in status_values + DEFAULT_STATUSES:
            if name not in merged:
                merged.append(name)
        self._statuses = merged

    def _extract_validations(self, wb, ws) -> Dict[str, List[str]]:
        validations: Dict[str, List[str]] = {}
        dv_list = getattr(ws, "data_validations", None)
        if not dv_list:
            return validations

        for dv in getattr(dv_list, "dataValidation", []) or []:
            if dv.type != "list":
                continue
            values = self._resolve_validation_values(wb, ws, dv)
            if not values:
                continue
            try:
                ranges = list(dv.ranges)
            except TypeError:
                ranges = []
            for cell_range in ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
                for col_idx in range(min_col, max_col + 1):
                    header = ws.cell(row=1, column=col_idx).value
                    if isinstance(header, str) and header in REQUIRED_COLUMNS:
                        validations[header] = list(values)
        return validations

    def _resolve_validation_values(self, wb, ws, dv) -> List[str]:
        formula = (dv.formula1 or "").strip()
        if not formula:
            return []
        if formula.startswith('"') and formula.endswith('"'):
            content = formula[1:-1]
            parts = [p.replace('""', '"').strip() for p in content.split(",")]
            return [p for p in parts if p]
        if formula.startswith("="):
            target = formula[1:]
            sheet_name = ws.title
            if "!" in target:
                sheet_part, range_part = target.split("!", 1)
                sheet_name = sheet_part.strip()
                if sheet_name.startswith("'") and sheet_name.endswith("'"):
                    sheet_name = sheet_name[1:-1].replace("''", "'")
            else:
                range_part = target
            range_part = range_part.strip()
            try:
                target_ws = wb[sheet_name]
            except KeyError:
                return []
            try:
                min_col, min_row, max_col, max_row = range_boundaries(range_part)
            except ValueError:
                return []

            try:
                used_min_col, used_min_row, used_max_col, used_max_row = range_boundaries(
                    target_ws.calculate_dimension()
                )
            except ValueError:
                used_min_col = used_max_col = min_col
                used_min_row = used_max_row = min_row

            min_col = max(min_col, used_min_col)
            max_col = min(max_col, used_max_col)
            min_row = max(min_row, used_min_row)
            max_row = min(max_row, used_max_row)

            if min_col > max_col or min_row > max_row:
                return []

            width = max_col - min_col + 1
            height = max_row - min_row + 1
            if width * height > MAX_VALIDATION_CELLS:
                max_height = max(1, MAX_VALIDATION_CELLS // width)
                max_row = min(max_row, min_row + max_height - 1)

            values: List[str] = []
            seen = set()
            for row in target_ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            ):
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value).strip()
                    if not text or text in seen:
                        continue
                    seen.add(text)
                    values.append(text)
            return values
        return []

    def _to_excel_value(self, col_name: str, value: Any):
        if value is None or value is pd.NA:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass

        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return value

        if col_name == "No":
            try:
                return int(value)
            except Exception:
                return value

        if col_name == "期限":
            if isinstance(value, str):
                parsed = _from_iso_date_str(value)
                return None if parsed is pd.NaT else parsed

        if isinstance(value, str):
            text = value.strip()
            return text if text else None

        return value

    def _build_validation_formula(self, values: List[str]) -> str:
        escaped = [v.replace('"', '""') for v in values]
        return '"' + ",".join(escaped) + '"'

    def _ensure_unique_no(self, no_value: int):
        known = set(
            self._df["No"].dropna().astype(int).tolist()
        )
        if int(no_value) in known:
            raise ValueError(f"No={no_value} は既に存在します。")

    def _next_no(self) -> int:
        if self._df.empty or self._df["No"].dropna().empty:
            return 1
        return int(self._df["No"].dropna().astype(int).max()) + 1

    def get_tasks(self) -> List[Dict[str, Any]]:
        with self._lock:
            return [self._format_row(self._df.loc[idx]) for idx in self._df.index]

    def get_statuses(self) -> List[str]:
        with self._lock:
            return list(self._statuses)

    def get_validations(self) -> Dict[str, List[str]]:
        with self._lock:
            return {k: list(v) for k, v in self._validations.items()}

    def set_validations(self, mapping: Dict[str, List[Any]]):
        with self._lock:
            cleaned: Dict[str, List[str]] = {}
            for col in REQUIRED_COLUMNS:
                if col == "No":
                    continue
                raw_values = mapping.get(col)
                if not raw_values:
                    continue
                values: List[str] = []
                for value in raw_values:
                    text = str(value or "").strip()
                    if text and text not in values:
                        values.append(text)
                if values:
                    cleaned[col] = values
            self._validations = cleaned
            if cleaned.get("ステータス"):
                base = list(cleaned["ステータス"])
                extras = [
                    s
                    for s in self._df["ステータス"].dropna().astype(str).tolist()
                    if s and s not in base
                ]
                self._statuses = base + extras
            else:
                self._rebuild_statuses_from_df()

    def _format_row(self, row: pd.Series) -> Dict[str, Any]:
        return {
            "No": None if pd.isna(row["No"]) else int(row["No"]),
            "ステータス": "" if pd.isna(row["ステータス"]) else str(row["ステータス"]),
            "タスク": "" if pd.isna(row["タスク"]) else str(row["タスク"]),
            "担当者": "" if pd.isna(row["担当者"]) else str(row["担当者"]),
            "優先度": _format_priority(row["優先度"]),
            "期限": _to_iso_date_str(row["期限"]),
            "備考": "" if pd.isna(row["備考"]) else str(row["備考"]),
        }

    def _ensure_status_registered(self, name: str):
        name = name.strip()
        if name and name not in self._statuses:
            self._statuses.append(name)

    def add_task(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        with self._lock:
            task = dict(payload)
            raw_no = task.get("No")
            if raw_no in (None, "", pd.NA):
                no_value = self._next_no()
            else:
                no_value = int(raw_no)
                self._ensure_unique_no(no_value)

            status = str(task.get("ステータス", "") or "").strip()
            title = str(task.get("タスク", "") or "").strip()
            assignee = str(task.get("担当者", "") or "").strip()
            notes = str(task.get("備考", "") or "")
            due = _from_iso_date_str(task.get("期限", ""))
            priority = _normalize_priority(task.get("優先度"))

            row = {
                "No": int(no_value),
                "ステータス": status,
                "タスク": title,
                "担当者": assignee,
                "優先度": priority,
                "期限": due,
                "備考": notes,
            }

            self._df = pd.concat([self._df, pd.DataFrame([row])], ignore_index=True)
            self._ensure_status_registered(status)
            return self._format_row(self._df.iloc[-1])

    def update_task(self, no_value: int, patch: Dict[str, Any]) -> Dict[str, Any]:
        with self._lock:
            idx = self._df.index[self._df["No"].astype("Int64") == int(no_value)]
            if len(idx) == 0:
                raise KeyError(f"No={no_value} は存在しません。")

            i = idx[0]
            if "ステータス" in patch:
                status = str(patch["ステータス"] or "").strip()
                self._ensure_status_registered(status)
                self._df.at[i, "ステータス"] = status
            if "タスク" in patch:
                self._df.at[i, "タスク"] = str(patch["タスク"] or "").strip()
            if "担当者" in patch:
                self._df.at[i, "担当者"] = str(patch["担当者"] or "").strip()
            if "優先度" in patch:
                self._df.at[i, "優先度"] = _normalize_priority(patch["優先度"])
            if "期限" in patch:
                self._df.at[i, "期限"] = _from_iso_date_str(patch["期限"])
            if "備考" in patch:
                self._df.at[i, "備考"] = str(patch["備考"] or "")

            return self._format_row(self._df.loc[i])

    def move_task(self, no_value: int, new_status: str) -> Dict[str, Any]:
        return self.update_task(int(no_value), {"ステータス": new_status})

    def delete_task(self, no_value: int) -> bool:
        with self._lock:
            idx = self._df.index[self._df["No"].astype("Int64") == int(no_value)]
            if len(idx) == 0:
                return False
            self._df = self._df.drop(index=idx).reset_index(drop=True)
            return True

    def save_excel(self) -> str:
        with self._lock:
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            tmp_path = self.excel_path.with_name(
                f"{self.excel_path.stem}.tmp_{ts}{self.excel_path.suffix}"
            )
            backup_path = self.excel_path.with_name(
                f"{self.excel_path.stem}.bak_{ts}{self.excel_path.suffix}"
            )

            df = self._df.copy()
            try:
                if self.excel_path.exists():
                    wb = load_workbook(self.excel_path)
                else:
                    wb = Workbook()
                if not self._sheet_name:
                    self._sheet_name = wb.sheetnames[0]
                if self._sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=self._sheet_name)
                else:
                    ws = wb[self._sheet_name]

                ws.delete_rows(1, ws.max_row)
                ws.append(REQUIRED_COLUMNS)
                for row in df.itertuples(index=False, name=None):
                    values: List[Any] = []
                    for col_name, value in zip(REQUIRED_COLUMNS, row):
                        values.append(self._to_excel_value(col_name, value))
                    ws.append(values)

                if ws.data_validations is not None:
                    ws.data_validations.dataValidation = []

                for col_name, values in self._validations.items():
                    if col_name not in REQUIRED_COLUMNS or not values:
                        continue
                    try:
                        idx = REQUIRED_COLUMNS.index(col_name) + 1
                    except ValueError:
                        continue
                    col_letter = get_column_letter(idx)
                    formula = self._build_validation_formula(values)
                    dv = DataValidation(
                        type="list",
                        formula1=formula,
                        allow_blank=True,
                        showDropDown=True,
                    )
                    dv.add(f"${col_letter}$2:${col_letter}$1048576")
                    ws.add_data_validation(dv)

                wb.save(tmp_path)
                if self.excel_path.exists():
                    shutil.copy2(self.excel_path, backup_path)
                os.replace(tmp_path, self.excel_path)
            except Exception:
                if tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except Exception:
                        pass
                raise
            return str(self.excel_path.resolve())


class JsApi:
    def __init__(self, store: TaskStore):
        self.store = store

    def get_tasks(self) -> List[Dict[str, Any]]:
            return self.store.get_tasks()

    def get_statuses(self) -> List[str]:
        return self.store.get_statuses()

    def get_validations(self) -> Dict[str, List[str]]:
        return self.store.get_validations()

    def update_validations(self, payload: Any) -> Dict[str, Any]:
        data = json.loads(payload) if isinstance(payload, str) else payload
        self.store.set_validations(data or {})
        return {
            "ok": True,
            "validations": self.store.get_validations(),
            "statuses": self.store.get_statuses(),
        }

    def add_task(self, payload: Any) -> Dict[str, Any]:
        data = json.loads(payload) if isinstance(payload, str) else payload
        return self.store.add_task(data)

    def update_task(self, no_value: Any, patch: Any) -> Dict[str, Any]:
        patch_data = json.loads(patch) if isinstance(patch, str) else patch
        return self.store.update_task(int(no_value), patch_data)

    def move_task(self, no_value: Any, new_status: str) -> Dict[str, Any]:
        return self.store.move_task(int(no_value), new_status)

    def delete_task(self, no_value: Any) -> bool:
        return self.store.delete_task(int(no_value))

    def save_excel(self) -> str:
        return self.store.save_excel()

    def reload_from_excel(self) -> Dict[str, Any]:
        self.store.load_excel()
        return {
            "ok": True,
            "tasks": self.store.get_tasks(),
            "statuses": self.store.get_statuses(),
            "validations": self.store.get_validations(),
        }


def main():
    parser = argparse.ArgumentParser(description="Excel Kanban backend")
    parser.add_argument("--excel", default="./task.xlsx")
    parser.add_argument("--html", default="./index.html")
    parser.add_argument("--title", default="タスク・ボード")
    parser.add_argument("--width", type=int, default=1280)
    parser.add_argument("--height", type=int, default=800)
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    html_path = Path(args.html).expanduser().resolve()
    if not html_path.exists():
        raise FileNotFoundError(f"HTML が見つかりません: {html_path}")

    store = TaskStore(excel_path)
    api = JsApi(store)

    window = webview.create_window(
        title=args.title,
        url=html_path.as_uri(),
        width=args.width,
        height=args.height,
        js_api=api,
    )

    webview.start(
        func=None,
        debug=args.debug,
        http_server=False,
        gui="edgechromium",
        private_mode=False,
    )


if __name__ == "__main__":
    main()
