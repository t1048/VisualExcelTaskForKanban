# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import os
import shutil
import threading
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import datetime as dt

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
import webview
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from watchdog.events import FileSystemEventHandler
    from watchdog.observers import Observer
    from watchdog.observers.polling import PollingObserver
except ImportError:  # pragma: no cover - optional dependency guard
    FileSystemEventHandler = object  # type: ignore
    Observer = None  # type: ignore
    PollingObserver = None  # type: ignore


TASK_COLUMNS = [
    "ステータス",
    "大分類",
    "中分類",
    "タスク",
    "担当者",
    "優先度",
    "期限",
    "備考",
]
DEFAULT_STATUSES = ["未着手", "進行中", "完了", "保留"]
DEFAULT_PRIORITY_LEVELS = ["高", "中", "低"]
DEFAULT_VALIDATIONS: Dict[str, List[str]] = {
    "ステータス": list(DEFAULT_STATUSES),
    "優先度": list(DEFAULT_PRIORITY_LEVELS),
}


def _load_exec_options(path: Path) -> Dict[str, Any]:
    try:
        with path.open("r", encoding="utf-8") as fp:
            data = json.load(fp)
        return data if isinstance(data, dict) else {}
    except FileNotFoundError:
        return {}
    except json.JSONDecodeError:
        return {}


def _save_exec_options(path: Path, options: Dict[str, Any]) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as fp:
            json.dump(options, fp, ensure_ascii=False, indent=2)
    except OSError:
        print(f"[kanban] 設定ファイルを保存できませんでした: {path}")


def _open_option_dialog(initial_options: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    root = tk.Tk()
    root.title("Excel Kanban 起動設定")
    root.resizable(False, False)

    result: Dict[str, Any] = {}
    cancelled = {"value": False}

    def on_cancel():
        cancelled["value"] = True
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_cancel)

    main_frame = ttk.Frame(root, padding=10)
    main_frame.grid(row=0, column=0, sticky="nsew")
    main_frame.columnconfigure(1, weight=1)

    string_fields = {
        "excel": tk.StringVar(value=str(initial_options.get("excel", ""))),
        "html": tk.StringVar(value=str(initial_options.get("html", ""))),
        "title": tk.StringVar(value=str(initial_options.get("title", ""))),
        "sheet": tk.StringVar(value=str(initial_options.get("sheet", "") or "")),
        "width": tk.StringVar(value=str(initial_options.get("width", ""))),
        "height": tk.StringVar(value=str(initial_options.get("height", ""))),
        "watch_interval": tk.StringVar(value=str(initial_options.get("watch_interval", ""))),
        "watch_debounce": tk.StringVar(value=str(initial_options.get("watch_debounce", ""))),
    }

    bool_fields = {
        "debug": tk.BooleanVar(value=bool(initial_options.get("debug", False))),
        "no_watch": tk.BooleanVar(value=bool(initial_options.get("no_watch", False))),
        "watch_polling": tk.BooleanVar(value=bool(initial_options.get("watch_polling", False))),
    }

    def build_file_row(row: int, label: str, field_key: str, filetypes: Tuple[Tuple[str, str], ...]):
        ttk.Label(main_frame, text=label).grid(row=row, column=0, sticky="w", pady=2)
        entry = ttk.Entry(main_frame, textvariable=string_fields[field_key], width=50)
        entry.grid(row=row, column=1, sticky="we", pady=2)

        def browse():
            initial = string_fields[field_key].get() or ""
            initial_dir = os.path.dirname(initial) if initial else os.getcwd()
            selected = filedialog.askopenfilename(
                title=f"{label} を選択",
                initialdir=initial_dir,
                filetypes=filetypes,
            )
            if selected:
                string_fields[field_key].set(selected)

        ttk.Button(main_frame, text="参照", command=browse).grid(row=row, column=2, padx=(4, 0))

    build_file_row(0, "Excel ファイル", "excel", (("Excel", "*.xlsx"), ("すべて", "*.*")))
    build_file_row(1, "HTML ファイル", "html", (("HTML", "*.html"), ("すべて", "*.*")))

    ttk.Label(main_frame, text="ウィンドウタイトル").grid(row=2, column=0, sticky="w", pady=2)
    ttk.Entry(main_frame, textvariable=string_fields["title"], width=50).grid(
        row=2, column=1, columnspan=2, sticky="we", pady=2
    )

    ttk.Label(main_frame, text="ウィンドウサイズ (幅 x 高さ)").grid(
        row=3, column=0, sticky="w", pady=2
    )
    size_frame = ttk.Frame(main_frame)
    size_frame.grid(row=3, column=1, columnspan=2, sticky="w", pady=2)
    ttk.Entry(size_frame, textvariable=string_fields["width"], width=10).grid(row=0, column=0)
    ttk.Label(size_frame, text="x").grid(row=0, column=1, padx=4)
    ttk.Entry(size_frame, textvariable=string_fields["height"], width=10).grid(row=0, column=2)

    ttk.Label(main_frame, text="シート名").grid(row=4, column=0, sticky="w", pady=2)
    ttk.Entry(main_frame, textvariable=string_fields["sheet"], width=20).grid(
        row=4, column=1, columnspan=2, sticky="we", pady=2
    )

    ttk.Label(main_frame, text="監視設定").grid(row=5, column=0, sticky="w", pady=(8, 2))
    checks_frame = ttk.Frame(main_frame)
    checks_frame.grid(row=5, column=1, columnspan=2, sticky="w", pady=(8, 2))
    ttk.Checkbutton(
        checks_frame,
        text="デバッグモード",
        variable=bool_fields["debug"],
    ).grid(row=0, column=0, sticky="w")
    ttk.Checkbutton(
        checks_frame,
        text="ファイル監視を無効化",
        variable=bool_fields["no_watch"],
    ).grid(row=0, column=1, sticky="w", padx=(10, 0))
    ttk.Checkbutton(
        checks_frame,
        text="PollingObserver を使用",
        variable=bool_fields["watch_polling"],
    ).grid(row=0, column=2, sticky="w", padx=(10, 0))

    ttk.Label(main_frame, text="監視間隔 (秒)").grid(row=6, column=0, sticky="w", pady=2)
    ttk.Entry(main_frame, textvariable=string_fields["watch_interval"], width=10).grid(
        row=6, column=1, sticky="w", pady=2
    )

    ttk.Label(main_frame, text="監視ディレイ (秒)").grid(row=7, column=0, sticky="w", pady=2)
    ttk.Entry(main_frame, textvariable=string_fields["watch_debounce"], width=10).grid(
        row=7, column=1, sticky="w", pady=2
    )

    button_frame = ttk.Frame(main_frame)
    button_frame.grid(row=8, column=0, columnspan=3, pady=(12, 0))

    def on_submit():
        try:
            width_value = int(string_fields["width"].get())
            height_value = int(string_fields["height"].get())
        except ValueError:
            messagebox.showerror("入力エラー", "幅と高さには整数を入力してください。")
            return

        try:
            watch_interval_value = float(string_fields["watch_interval"].get())
            watch_debounce_value = float(string_fields["watch_debounce"].get())
        except ValueError:
            messagebox.showerror("入力エラー", "監視間隔と監視ディレイには数値を入力してください。")
            return

        result.update(
            {
                "excel": string_fields["excel"].get().strip(),
                "html": string_fields["html"].get().strip(),
                "title": string_fields["title"].get().strip(),
                "width": width_value,
                "height": height_value,
                "sheet": string_fields["sheet"].get().strip(),
                "debug": bool_fields["debug"].get(),
                "no_watch": bool_fields["no_watch"].get(),
                "watch_polling": bool_fields["watch_polling"].get(),
                "watch_interval": watch_interval_value,
                "watch_debounce": watch_debounce_value,
            }
        )
        root.destroy()

    ttk.Button(button_frame, text="キャンセル", command=on_cancel).grid(row=0, column=0, padx=5)
    ttk.Button(button_frame, text="起動", command=on_submit).grid(row=0, column=1, padx=5)

    root.mainloop()

    if cancelled["value"]:
        return None

    result["sheet"] = result.get("sheet") or ""
    return result


def _to_iso_date_str(value) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (dt.date, dt.datetime, pd.Timestamp)):
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    try:
        return pd.to_datetime(str(value)).strftime("%Y-%m-%d")
    except Exception:
        return str(value)


def _from_iso_date_str(value: str):
    value = (value or "").strip()
    if not value:
        return pd.NaT
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return pd.NaT


def _normalize_priority(value: Any):
    if value is None or value is pd.NA:
        return pd.NA

    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return pd.NA
        value = trimmed

    if pd.isna(value):
        return pd.NA

    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)

    if pd.isna(num):
        return pd.NA

    if float(num).is_integer():
        return int(num)
    return float(num)


def _format_priority(value: Any):
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float)) and not pd.isna(value):
        if isinstance(value, float) and float(value).is_integer():
            return int(value)
        return value
    return str(value)


class TaskStore:
    def __init__(self, excel_path: Path, sheet_name: str | None = None):
        self.excel_path = excel_path
        self._lock = threading.RLock()
        self._df = pd.DataFrame(columns=TASK_COLUMNS)
        self._statuses: List[str] = list(DEFAULT_STATUSES)
        self._requested_sheet_name: str | None = (
            sheet_name.strip() if isinstance(sheet_name, str) and sheet_name.strip() else None
        )
        self._sheet_name: str | None = self._requested_sheet_name
        self._validations: Dict[str, List[str]] = {
            key: list(values) for key, values in DEFAULT_VALIDATIONS.items()
        }
        self._last_saved_at: Optional[dt.datetime] = None
        self._last_saved_mtime: Optional[float] = None
        self._last_loaded_mtime: Optional[float] = None
        self.load_excel()

    def load_excel(self):
        with self._lock:
            requested_sheet = self._requested_sheet_name
            if not self.excel_path.exists():
                df = pd.DataFrame(columns=TASK_COLUMNS)
                df.to_excel(
                    self.excel_path,
                    index=False,
                    sheet_name=requested_sheet or "Sheet1",
                )

            wb = load_workbook(self.excel_path, data_only=False)
            if requested_sheet:
                if requested_sheet in wb.sheetnames:
                    sheet_name = requested_sheet
                else:
                    ws = wb.create_sheet(title=requested_sheet)
                    ws.append(TASK_COLUMNS)
                    wb.save(self.excel_path)
                    sheet_name = requested_sheet
            else:
                sheet_name = wb.sheetnames[0]
            ws = wb[sheet_name]
            self._sheet_name = sheet_name
            extracted = self._extract_validations(wb, ws)
            validations: Dict[str, List[str]] = {
                key: list(values) for key, values in DEFAULT_VALIDATIONS.items()
            }
            for column, values in extracted.items():
                if not values:
                    continue
                validations[column] = list(values)
            self._validations = validations

            df = pd.read_excel(
                self.excel_path,
                sheet_name=sheet_name,
                engine="openpyxl",
            )

            if "No" in df.columns:
                df = df.drop(columns=["No"])

            for col in TASK_COLUMNS:
                if col not in df.columns:
                    df[col] = pd.NA

            df = df[TASK_COLUMNS].copy()

            df = df.dropna(how="all", subset=TASK_COLUMNS).reset_index(drop=True)

            df["タスク"] = df["タスク"].apply(
                lambda v: "" if pd.isna(v) else str(v).strip()
            )
            df = df[df["タスク"] != ""].reset_index(drop=True)

            if "期限" in df.columns:
                df["期限"] = pd.to_datetime(df["期限"], errors="coerce").dt.date

            if self._validations.get("ステータス"):
                base = list(self._validations["ステータス"])
                extras = [
                    s
                    for s in df["ステータス"].dropna().astype(str).tolist()
                    if s and s not in base
                ]
                self._statuses = base + extras
            else:
                self._rebuild_statuses_from_df(df)

            self._df = df
            self._last_loaded_mtime = self._get_file_mtime()
            if self._last_saved_mtime is None:
                self._last_saved_mtime = self._last_loaded_mtime

    def _get_file_mtime(self) -> Optional[float]:
        try:
            return self.excel_path.stat().st_mtime
        except FileNotFoundError:
            return None

    def get_last_saved_markers(self) -> Tuple[Optional[dt.datetime], Optional[float]]:
        with self._lock:
            return self._last_saved_at, self._last_saved_mtime

    def _rebuild_statuses_from_df(self, df: pd.DataFrame | None = None):
        if df is None:
            df = self._df
        status_values = [
            s for s in df["ステータス"].dropna().astype(str).tolist() if s
        ]
        merged: List[str] = []
        for name in status_values + DEFAULT_STATUSES:
            if name not in merged:
                merged.append(name)
        self._statuses = merged

    def _extract_validations(self, wb, ws) -> Dict[str, List[str]]:
        validations: Dict[str, List[str]] = {}
        dv_list = getattr(ws, "data_validations", None)
        if not dv_list:
            return validations

        for dv in getattr(dv_list, "dataValidation", []) or []:
            if dv.type != "list":
                continue
            values = self._resolve_validation_values(wb, ws, dv)
            if not values:
                continue
            try:
                ranges = list(dv.ranges)
            except TypeError:
                ranges = []
            for cell_range in ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
                for col_idx in range(min_col, max_col + 1):
                    header = ws.cell(row=1, column=col_idx).value
                    if isinstance(header, str) and header in TASK_COLUMNS:
                        validations[header] = list(values)
        return validations

    def _resolve_validation_values(self, wb, ws, dv) -> List[str]:
        formula = (dv.formula1 or "").strip()
        if not formula:
            return []
        if formula.startswith('"') and formula.endswith('"'):
            content = formula[1:-1]
            parts = [p.replace('""', '"').strip() for p in content.split(",")]
            return [p for p in parts if p]
        if formula.startswith("="):
            target = formula[1:]
            sheet_name = ws.title
            if "!" in target:
                sheet_part, range_part = target.split("!", 1)
                sheet_name = sheet_part.strip()
                if sheet_name.startswith("'") and sheet_name.endswith("'"):
                    sheet_name = sheet_name[1:-1].replace("''", "'")
            else:
                range_part = target
            range_part = range_part.strip()
            try:
                target_ws = wb[sheet_name]
            except KeyError:
                return []
            values: List[str] = []
            for row in target_ws[range_part]:
                cells = row if isinstance(row, (list, tuple)) else (row,)
                for cell in cells:
                    if cell.value is None:
                        continue
                    text = str(cell.value).strip()
                    if not text:
                        continue
                    if text not in values:
                        values.append(text)
            return values
        return []

    def _to_excel_value(self, col_name: str, value: Any):
        if value is None or value is pd.NA:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass

        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime()
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return value

        if col_name == "期限":
            if isinstance(value, str):
                parsed = _from_iso_date_str(value)
                return None if parsed is pd.NaT else parsed

        if isinstance(value, str):
            text = value.strip()
            if col_name == "備考" and text:
                # Excel セル内での改行は LF ("\n") で扱われるが、入力元によっては
                # CRLF ("\r\n") や単独の CR ("\r") が混在することがある。
                # CR を維持したまま保存すると Excel 上では 1 回の改行が複数回に
                # 表示されてしまうため、LF のみに正規化して書き込む。
                text = text.replace("\r\n", "\n").replace("\r", "\n")
            return text if text else None

        return value

    def _build_validation_formula(self, values: List[str]) -> str:
        escaped = [v.replace('"', '""') for v in values]
        return '"' + ",".join(escaped) + '"'

    def get_tasks(self) -> List[Dict[str, Any]]:
        with self._lock:
            return [self._format_row(i, self._df.iloc[i]) for i in range(len(self._df))]

    def get_statuses(self) -> List[str]:
        with self._lock:
            return list(self._statuses)

    def get_validations(self) -> Dict[str, List[str]]:
        with self._lock:
            return {k: list(v) for k, v in self._validations.items()}

    def set_validations(self, mapping: Dict[str, List[Any]]):
        with self._lock:
            cleaned: Dict[str, List[str]] = {}
            for col in TASK_COLUMNS:
                raw_values = mapping.get(col)
                if not raw_values:
                    continue
                values: List[str] = []
                for value in raw_values:
                    text = str(value or "").strip()
                    if text and text not in values:
                        values.append(text)
                if values:
                    cleaned[col] = values
            merged: Dict[str, List[str]] = {
                key: list(values) for key, values in DEFAULT_VALIDATIONS.items()
            }
            merged.update(cleaned)
            self._validations = merged
            if self._validations.get("ステータス"):
                base = list(self._validations["ステータス"])
                extras = [
                    s
                    for s in self._df["ステータス"].dropna().astype(str).tolist()
                    if s and s not in base
                ]
                self._statuses = base + extras
            else:
                self._rebuild_statuses_from_df()

    def _format_row(self, idx: int, row: pd.Series) -> Dict[str, Any]:
        return {
            "No": int(idx + 1),
            "ステータス": "" if pd.isna(row["ステータス"]) else str(row["ステータス"]),
            "大分類": "" if pd.isna(row["大分類"]) else str(row["大分類"]),
            "中分類": "" if pd.isna(row["中分類"]) else str(row["中分類"]),
            "タスク": "" if pd.isna(row["タスク"]) else str(row["タスク"]),
            "担当者": "" if pd.isna(row["担当者"]) else str(row["担当者"]),
            "優先度": _format_priority(row["優先度"]),
            "期限": _to_iso_date_str(row["期限"]),
            "備考": "" if pd.isna(row["備考"]) else str(row["備考"]),
        }

    def _resolve_row_index(self, no_value: int) -> int:
        try:
            no = int(no_value)
        except (TypeError, ValueError):
            raise KeyError(f"No={no_value} は存在しません。")
        if no <= 0 or no > len(self._df):
            raise KeyError(f"No={no_value} は存在しません。")
        return no - 1

    def _ensure_status_registered(self, name: str):
        name = name.strip()
        if name and name not in self._statuses:
            self._statuses.append(name)

    def add_task(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        with self._lock:
            task = dict(payload)

            status = str(task.get("ステータス", "") or "").strip()
            major = str(task.get("大分類", "") or "").strip()
            minor = str(task.get("中分類", "") or "").strip()
            title = str(task.get("タスク", "") or "").strip()
            assignee = str(task.get("担当者", "") or "").strip()
            notes = str(task.get("備考", "") or "")
            due = _from_iso_date_str(task.get("期限", ""))
            priority = _normalize_priority(task.get("優先度"))

            if not title:
                raise ValueError("タスクは必須項目です。")

            row = {
                "ステータス": status,
                "大分類": major,
                "中分類": minor,
                "タスク": title,
                "担当者": assignee,
                "優先度": priority,
                "期限": due,
                "備考": notes,
            }

            new_index = len(self._df)
            self._df.loc[new_index] = row
            self._ensure_status_registered(status)
            return self._format_row(new_index, self._df.iloc[new_index])

    def update_task(self, no_value: int, patch: Dict[str, Any]) -> Dict[str, Any]:
        with self._lock:
            i = self._resolve_row_index(int(no_value))
            row_index = self._df.index[i]

            updates: Dict[str, Any] = {}

            if "ステータス" in patch:
                updates["ステータス"] = str(patch["ステータス"] or "").strip()
            if "大分類" in patch:
                updates["大分類"] = str(patch["大分類"] or "").strip()
            if "中分類" in patch:
                updates["中分類"] = str(patch["中分類"] or "").strip()
            if "タスク" in patch:
                title = str(patch["タスク"] or "").strip()
                if not title:
                    raise ValueError("タスクは必須項目です。")
                updates["タスク"] = title
            if "担当者" in patch:
                updates["担当者"] = str(patch["担当者"] or "").strip()
            if "優先度" in patch:
                updates["優先度"] = _normalize_priority(patch["優先度"])
            if "期限" in patch:
                updates["期限"] = _from_iso_date_str(patch["期限"])
            if "備考" in patch:
                updates["備考"] = str(patch["備考"] or "")

            if "ステータス" in updates:
                self._ensure_status_registered(updates["ステータス"])

            for column, value in updates.items():
                self._df.at[row_index, column] = value

            return self._format_row(i, self._df.loc[row_index])

    def move_task(self, no_value: int, new_status: str) -> Dict[str, Any]:
        return self.update_task(int(no_value), {"ステータス": new_status})

    def delete_task(self, no_value: int) -> bool:
        with self._lock:
            try:
                idx = self._resolve_row_index(int(no_value))
            except KeyError:
                return False
            self._df = self._df.drop(index=self._df.index[idx]).reset_index(drop=True)
            return True

    def save_excel(self) -> str:
        with self._lock:
            ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = self.excel_path.with_name(
                f"{self.excel_path.stem}.bak_{ts}{self.excel_path.suffix}"
            )
            tmp_path = self.excel_path.with_name(
                f"{self.excel_path.stem}.tmp_{ts}{self.excel_path.suffix}"
            )

            df = self._df.copy()
            try:
                if self.excel_path.exists():
                    wb = load_workbook(self.excel_path)
                else:
                    wb = Workbook()
                if not self._sheet_name:
                    self._sheet_name = wb.sheetnames[0]
                if self._sheet_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=self._sheet_name)
                else:
                    ws = wb[self._sheet_name]

                ws.delete_rows(1, ws.max_row)
                ws.append(TASK_COLUMNS)
                for row in df.itertuples(index=False, name=None):
                    values: List[Any] = []
                    for col_name, value in zip(TASK_COLUMNS, row):
                        values.append(self._to_excel_value(col_name, value))
                    ws.append(values)

                try:
                    due_col_idx = TASK_COLUMNS.index("期限") + 1
                except ValueError:
                    due_col_idx = None

                if due_col_idx is not None and ws.max_row >= 2:
                    for cell in ws.iter_rows(
                        min_row=2,
                        max_row=ws.max_row,
                        min_col=due_col_idx,
                        max_col=due_col_idx,
                    ):
                        for target in cell:
                            target.number_format = "yyyy/mm/dd"

                if ws.data_validations is not None:
                    ws.data_validations.dataValidation = []

                for col_name, values in self._validations.items():
                    if col_name not in TASK_COLUMNS or not values:
                        continue
                    try:
                        idx = TASK_COLUMNS.index(col_name) + 1
                    except ValueError:
                        continue
                    col_letter = get_column_letter(idx)
                    formula = self._build_validation_formula(values)
                    dv = DataValidation(
                        type="list",
                        formula1=formula,
                        allow_blank=True,
                        showDropDown=False,
                    )
                    dv.add(f"${col_letter}$2:${col_letter}$1048576")
                    ws.add_data_validation(dv)

                wb.save(tmp_path)
                os.replace(tmp_path, self.excel_path)
                shutil.copy2(self.excel_path, backup_path)
                self._last_saved_at = dt.datetime.now()
                self._last_saved_mtime = self._get_file_mtime()
                self._last_loaded_mtime = self._last_saved_mtime
            except Exception:
                if tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except Exception:
                        pass
                raise
            return str(self.excel_path.resolve())


def push_excel_update(window, store: TaskStore):
    payload = {
        "tasks": store.get_tasks(),
        "statuses": store.get_statuses(),
        "validations": store.get_validations(),
    }
    json_payload = json.dumps(payload, ensure_ascii=False)
    script = (
        "if (window.__kanban_receive_update) {"
        f" window.__kanban_receive_update({json_payload});"
        " }"
    )
    try:
        window.evaluate_js(script)
    except Exception as exc:  # pragma: no cover - depends on runtime
        print(f"[kanban] Failed to push update to frontend: {exc}")


class ExcelFileWatcher(FileSystemEventHandler):
    def __init__(
        self,
        store: TaskStore,
        window,
        debounce_seconds: float = 2.0,
    ):
        super().__init__()
        self.store = store
        self.window = window
        self.debounce_seconds = max(0.0, float(debounce_seconds))
        try:
            self._target_path = store.excel_path.resolve()
        except Exception:
            self._target_path = store.excel_path
        self._last_notified_mtime: Optional[float] = None

    def on_modified(self, event):  # pragma: no cover - relies on filesystem events
        self._handle_event(event, [getattr(event, "src_path", None)])

    def on_created(self, event):  # pragma: no cover - relies on filesystem events
        self._handle_event(event, [getattr(event, "src_path", None)])

    def on_moved(self, event):  # pragma: no cover - relies on filesystem events
        paths = [getattr(event, "src_path", None), getattr(event, "dest_path", None)]
        self._handle_event(event, paths)

    def _handle_event(self, event, paths):
        if getattr(event, "is_directory", False):
            return
        for path_str in paths:
            if not path_str:
                continue
            if self._is_target(path_str):
                self._process_change()
                break

    def _is_target(self, path_str: str) -> bool:
        try:
            candidate = Path(path_str).resolve()
        except Exception:
            candidate = Path(path_str)
        try:
            return candidate == self._target_path or candidate.samefile(self._target_path)
        except Exception:
            return candidate == self._target_path

    def _process_change(self):
        now = dt.datetime.now()
        try:
            mtime = self.store.excel_path.stat().st_mtime
        except FileNotFoundError:
            return

        last_saved_at, last_saved_mtime = self.store.get_last_saved_markers()
        if (
            last_saved_at
            and last_saved_mtime is not None
            and abs(mtime - last_saved_mtime) < 0.5
            and (now - last_saved_at).total_seconds() < self.debounce_seconds
        ):
            return

        if self._last_notified_mtime is not None and abs(mtime - self._last_notified_mtime) < 0.5:
            return

        try:
            self.store.load_excel()
            push_excel_update(self.window, self.store)
            self._last_notified_mtime = mtime
            timestamp = now.strftime("%H:%M:%S")
            print(f"[kanban] Excel change detected ({timestamp}), board updated.")
        except Exception as exc:
            print(f"[kanban] Failed to handle Excel change: {exc}")


def start_excel_watcher(
    window,
    store: TaskStore,
    *,
    debounce_seconds: float = 2.0,
    use_polling: bool = False,
    poll_interval: float = 1.0,
):
    if Observer is None:
        print("[kanban] watchdog is not installed; file watching is disabled.")
        return None

    observer_cls = Observer
    if use_polling and PollingObserver is not None:
        observer_cls = PollingObserver

    try:
        if observer_cls is PollingObserver:
            observer = observer_cls(timeout=max(0.1, float(poll_interval)))
        else:
            observer = observer_cls()
    except TypeError:
        observer = observer_cls()

    handler = ExcelFileWatcher(store, window, debounce_seconds=debounce_seconds)
    watch_path = str(store.excel_path.parent)
    observer.schedule(handler, watch_path, recursive=False)
    observer.daemon = True
    observer.start()
    mode = "PollingObserver" if observer_cls is PollingObserver else "Observer"
    print(f"[kanban] Started Excel watcher using {mode} on {watch_path}.")
    return observer


class JsApi:
    def __init__(self, store: TaskStore):
        self.store = store

    def get_tasks(self) -> List[Dict[str, Any]]:
        return self.store.get_tasks()

    def get_statuses(self) -> List[str]:
        return self.store.get_statuses()

    def get_validations(self) -> Dict[str, List[str]]:
        return self.store.get_validations()

    def update_validations(self, payload: Any) -> Dict[str, Any]:
        data = json.loads(payload) if isinstance(payload, str) else payload
        self.store.set_validations(data or {})
        return {
            "ok": True,
            "validations": self.store.get_validations(),
            "statuses": self.store.get_statuses(),
        }

    def add_task(self, payload: Any) -> Dict[str, Any]:
        data = json.loads(payload) if isinstance(payload, str) else payload
        return self.store.add_task(data)

    def update_task(self, no_value: Any, patch: Any) -> Dict[str, Any]:
        patch_data = json.loads(patch) if isinstance(patch, str) else patch
        return self.store.update_task(int(no_value), patch_data)

    def move_task(self, no_value: Any, new_status: str) -> Dict[str, Any]:
        return self.store.move_task(int(no_value), new_status)

    def delete_task(self, no_value: Any) -> bool:
        return self.store.delete_task(int(no_value))

    def save_excel(self) -> str:
        return self.store.save_excel()

    def reload_from_excel(self) -> Dict[str, Any]:
        self.store.load_excel()
        return {
            "ok": True,
            "tasks": self.store.get_tasks(),
            "statuses": self.store.get_statuses(),
            "validations": self.store.get_validations(),
        }


def main():
    parser = argparse.ArgumentParser(description="Excel Kanban backend")
    parser.add_argument("--excel", default="./data/task.xlsx")
    parser.add_argument("--html", default="./frontend/pages/index.html")
    parser.add_argument("--title", default="タスク・ボード")
    parser.add_argument("--width", type=int, default=1280)
    parser.add_argument("--height", type=int, default=800)
    parser.add_argument("--debug", action="store_true")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Excel ファイルから読み込むシート名 (未指定時は先頭のシート)",
    )
    parser.add_argument(
        "--no-watch",
        action="store_true",
        help="Excel ファイルの変更監視を無効化します",
    )
    parser.add_argument(
        "--watch-polling",
        action="store_true",
        help="watchdog の PollingObserver を使用します (ネットワークドライブ等向け)",
    )
    parser.add_argument(
        "--watch-interval",
        type=float,
        default=1.0,
        help="PollingObserver 利用時の監視間隔（秒）",
    )
    parser.add_argument(
        "--watch-debounce",
        type=float,
        default=2.0,
        help="アプリ自身の保存直後に発生したイベントを無視する猶予時間（秒）",
    )
    parser.add_argument(
        "--config",
        default=None,
        help="起動オプションを保存する設定ファイルのパス (未指定時は ExecOption.json)",
    )
    parser.add_argument(
        "--no-gui",
        action="store_true",
        help="Tkinter による起動画面を表示せず、コマンドライン引数の値を使用します",
    )
    args = parser.parse_args()

    gui_fields = [
        "excel",
        "html",
        "title",
        "width",
        "height",
        "debug",
        "sheet",
        "no_watch",
        "watch_polling",
        "watch_interval",
        "watch_debounce",
    ]

    config_path = (
        Path(args.config).expanduser().resolve()
        if args.config
        else Path(__file__).with_name("ExecOption.json")
    )

    defaults = parser.parse_args([])
    current_options = {field: getattr(args, field) for field in gui_fields}
    stored_options = _load_exec_options(config_path)

    for field in gui_fields:
        if field in stored_options:
            default_value = getattr(defaults, field)
            current_value = getattr(args, field)
            if current_value == default_value:
                current_options[field] = stored_options[field]

    if not args.no_gui:
        selected_options = _open_option_dialog(current_options)
        if selected_options is None:
            print("[kanban] 起動がキャンセルされました。")
            return
        _save_exec_options(config_path, selected_options)
        for field in gui_fields:
            setattr(args, field, selected_options[field])
    else:
        for field in gui_fields:
            setattr(args, field, current_options[field])

    args.sheet = (args.sheet or "").strip() or None

    excel_path = Path(args.excel).expanduser().resolve()
    html_path = Path(args.html).expanduser().resolve()
    if not html_path.exists():
        raise FileNotFoundError(f"HTML が見つかりません: {html_path}")

    store = TaskStore(excel_path, sheet_name=args.sheet)
    api = JsApi(store)

    window = webview.create_window(
        title=args.title,
        url=html_path.as_uri(),
        width=args.width,
        height=args.height,
        js_api=api,
    )

    def bootstrap_file_watcher():  # pragma: no cover - runtime behaviour
        if args.no_watch:
            print("[kanban] File watcher disabled via --no-watch option.")
            return

        observer = start_excel_watcher(
            window,
            store,
            debounce_seconds=max(0.0, float(args.watch_debounce)),
            use_polling=args.watch_polling,
            poll_interval=float(args.watch_interval),
        )
        if observer is None:
            return

        def _stop_observer():
            try:
                observer.stop()
                observer.join(timeout=5)
                print("[kanban] Excel watcher stopped.")
            except Exception:
                pass

        events = getattr(window, "events", None)
        if events is not None and hasattr(events, "closed"):
            events.closed += _stop_observer

    webview.start(
        func=bootstrap_file_watcher,
        debug=args.debug,
        http_server=False,
        gui="edgechromium",
        private_mode=False,
    )


if __name__ == "__main__":
    main()
